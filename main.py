import time
import threading
import tkinter as tk
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
from openpyxl import Workbook, load_workbook

URL = "https://www.accuweather.com/pt/br/são-paulo/45881/current-weather/45881"
ARQUIVO_EXCEL = "dados_climaticos.xlsx"
INTERVALO = 60

navegador = None
executando = False

def iniciar_navegador():
    global navegador
    navegador = webdriver.Chrome()
    navegador.get(URL)

def capturar_dados():
    global navegador
    wait = WebDriverWait(navegador, 10)
    
    temp_atual = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.temp"))).text
    umidade_atual = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Umidade')]/following-sibling::div"))).text
    data_hora_atual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    return data_hora_atual, temp_atual, umidade_atual

def salvar_no_excel(dados):
    try:
        wb = load_workbook(ARQUIVO_EXCEL)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Data e Hora", "Temperatura", "Umidade"])
    
    ws.append(dados)
    wb.save(ARQUIVO_EXCEL)

def coletar_dados():
    global executando, navegador

    if navegador is None:
        iniciar_navegador()

    while executando:
        dados = capturar_dados()
        salvar_no_excel(dados)
        lbl_dados.config(text=f"Últimos Dados:\n{dados[0]}\nTemperatura: {dados[1]}\nUmidade: {dados[2]}")
        time.sleep(INTERVALO)

def iniciar_coleta():
    global executando

    if not executando:
        executando = True
        thread = threading.Thread(target=coletar_dados, daemon=True)
        thread.start()
        messagebox.showinfo("Iniciado", "Coleta de dados foi iniciada")

def parar_coleta():
    global executando, navegador

    if executando:
        executando = False
        if navegador:
            navegador.quit()
            navegador = None
        messagebox.showinfo("Parado", "Coleta de dados foi interrompida")

root = tk.Tk()
root.title("Coletor de Dados Climáticos")

btn_iniciar = tk.Button(root, text="iniciar programa", command=iniciar_coleta, font=("Arial", 12))
btn_iniciar.pack(pady=10)

btn_parar = tk.Button(root, text="parar programa", command=parar_coleta, font=("Arial", 12))
btn_parar.pack(pady=10)

lbl_dados = tk.Label(root, text="Últimos Dados: Nenhum ainda", font=("Arial", 12), justify="left")
lbl_dados.pack(pady=20)

root.mainloop()
